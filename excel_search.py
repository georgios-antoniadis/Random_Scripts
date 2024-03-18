import pandas as pd
from tkinter import filedialog
import time
from openpyxl import load_workbook

# read the Excel file into a pandas dataframe
#Dialog box to select storing location
target = filedialog.askopenfilename().replace("/","\\")
sanity_flag = True
while sanity_flag:
    if ".xlsx" not in target:
        print("Please pick a .xlsx file!")
        time.sleep(2)
        target = filedialog.askopenfilename().replace("/","\\")
    else:
        sanity_flag = False

workbook = load_workbook(filename=target)

# define a list of search terms
search_terms = [
"62465c0d-92f7-465b-951a-9ecc5f6c9526",
"3f4789a2-6edb-4899-bd12-81b79474cec3",
"209ea8a7-7dd3-42dc-9682-b2fffe4ce9d3",
"0ca1007c-ad05-47cb-8bc9-52f04b246822",
"f13d37f9-ee50-420c-a082-9991d262c08e",
"7430d0ab-73eb-422d-b3a8-82d07d2901d8",
"17f2455c-9fe7-4313-8ed4-c240e0aa6a92",
"6ab33719-6271-470d-ba2d-8a26b85c8e43",
"fc94d2d2-c66b-41b4-9b6f-a5602b07d3bc",
"7b592e83-c0e1-4d8f-84b1-2e70188bbf85",
"38ab3c60-b02b-4cfe-91bb-cc75a295726b",
"a4274b60-0ad9-4128-8b68-2176d163e312",
"318dd33b-14d2-4804-bb0e-4b779e9e152c",
"8ca6587c-b619-44ff-8961-1f50b6c12001",
"e1787e2a-546d-44c0-a97e-d92445d474a1",
"af0e3fb6-ab31-4de6-bf90-e8038476ca0e",
"17448af3-9211-45f1-9f21-38dd89e48ceb",
"b3854f68-9d10-4ddc-8c52-0e0b83bd46e7",
"32b61704-6853-404f-b62a-47c5b9641cb1",
"29f1c47e-6ac1-49c6-983f-1726b34001c9",
"b054cf68-71e8-4bc0-884a-29e08765636b",
"4ca4501b-4c20-41f7-9764-e17024ba894c",
"b43325a0-12ff-4e61-b7dd-4ea3c23f553a",
"ff2d93a7-16be-4364-a44b-045c96aa4e69",
"10b0a20c-ade9-4d2c-a46b-27c48d122f75",
"482bf260-7241-422b-a95f-32a971ff4962",
"421d0bfb-829f-4c27-8854-f336676c5d17",
"8a447516-d9ca-4566-89b9-266233164882",
"e30ec0d7-1d6e-48bc-b9c2-c485bdcc15f4",
"83f00bec-7e20-4d72-b60d-2b2d8bc6b302",
"eaee3d13-cf8c-4a84-8eba-7090867e4d02",
"2268cd44-7c4b-41da-a936-2fc94c388427",
"62952f92-6c01-474a-ac70-39427977a961",
"d678dd86-42c9-413d-8c6b-742e8d2b16f2",
"d4214b03-0b99-4349-8356-fdc6b3192d14",
"ac04f6a9-81f8-46f0-a623-92b4877ea9d5",
"7fc0c6f1-40b6-4b75-929e-76061c50e777",
"e1a56a9b-a1b5-4140-9dca-752d648e328f",
"1aec9245-b7e9-4bf1-8da4-11bd0cfe4190",
"d84e2917-a3b6-4447-a8d2-8b3f919b8e55",
"b9fe412d-403c-4a42-a3d1-ac66ebb237a6",
"7a154ebd-3132-49db-ba00-ab7a9a835002",
"0cef70ba-37fe-4402-a491-125b6cd9091f",
"de33634d-22e7-43bf-94e9-f0b0eaee1896",
"0f489caf-ced1-47cb-8789-36ef56f8c605",
"8bbbadd0-18c0-49fd-a14d-1719c4dd485a",
"014171b7-54a1-4bfa-bc8d-7fd83b16e15b",
"be03f103-908a-408c-8d0e-52ea891f86c6",
"98d75fd7-6523-4731-ab26-9ba9caa527e8",
"23cfd6e1-813c-4519-8ab5-fab2ddea2c01",
"20ebb951-fd7a-4113-8628-4d2b4d5dfc56",
"8800d419-f91a-440c-b4cd-37227268ce43",
"6b05aaee-6b12-4c71-a077-dd0f2cc9347e",
"4384a38b-5fb5-4fc7-8cfa-8bc337682abf",
"91fc326e-725b-444e-9fea-c8abd77d1c03",
"b6f2c6a0-790f-4133-bda9-67594ed43f52",
"f67800d3-0e54-46a5-85ee-4e12aa3c7147",
"7afd643c-6d07-43a4-8d29-33bcdbf0e99d",
"04a57d89-d969-4a39-a692-371dd091ebf6",
"310cc01d-d4fb-4128-9c85-976641a4d8bd",
"120923e3-181e-4453-b5aa-96ea46c57fea",
"c0561d07-8145-4ffd-92b9-584761af60c8",
"1e9817e5-906e-4652-a7d4-82dc67099b8c",
"90c7693e-9d40-4712-974a-7aacf401f702",
"36bf1544-a979-4ce4-b883-fcd278c03c08",
"5d25c0a2-294a-46e1-a7f6-8880a1b10123",
"8cfdf598-505e-4ef3-a953-63f65189a558",
"5fd81fc3-1942-4994-90a2-21dccf6d3102",
"a4c0e550-d2b0-4646-8dfc-c3a38b460a76",
]

# initialize a dictionary to store the search results
search_results = {}

# loop through all sheets in the workbook
for sheet in workbook:
    # loop through all cells in the sheet
    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            # loop through all search terms and check if they appear in the cell value
            for term in search_terms:
                if term in str(cell_value):
                    # increment the count of the search term in the search results dictionary
                    if term in search_results:
                        search_results[term] += 1
                    else:
                        search_results[term] = 1

terms_not_found_at_all = [search_term for search_term in search_terms if search_term not in list(search_results.keys())]

# print the search results
if search_results:
    for term, count in search_results.items():
        print(f'Term "{term}" appears {count} times in the Excel file.')
    if len(terms_not_found_at_all) != 0:
        print(f"Term(s) {terms_not_found_at_all} not found in the Excel file") 
else:
    print('None of the search terms appear in the Excel file.')